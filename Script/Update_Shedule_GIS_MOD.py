import os
import pandas as pd
from arcgis import GIS
from datetime import datetime
import numpy as np
from datetime import datetime
from copy import deepcopy
from pathlib import Path
from openpyxl import load_workbook
import traceback
import sys
# =================== VARIBALES LOCALES =====================
workspace 		= str(Path(__file__).parents[1])
excel       	= workspace+"\\EXCEL\\Schedule_sp.xlsx"
# excel       	= arcpy.GetParameterAsText(0)
name_general 	= "Update_Shedule"
logs 			= workspace+"\\LOGS\\Logs_%s.log"%name_general

portal      	= "https://andesgeogis.maps.arcgis.com"
user        	= "andesgeogis"
password    	= "Mercator1905#"
gis         	= GIS(portal,user,password)
id_survey     	= "e867d1ae2aed456b9cc37f3efa38e0c3"
id_shedule    	= "95916317bdfe4b44ab2eb53838dbe882" # Testting
# id_shedule    	= "5089a7be92ad4802af3e385091ad2515" # Produccion
now         	= datetime.now() 

# ======================== FUNCTIONS ===================================== 


def captureError(e,name):
	tb = sys.exc_info()[2]
	tbinfo = traceback.format_tb(tb)[0]
	write_log(0, 0, name)
	write_log(0, 0, str(e))	
	write_log(0, 0, str(tbinfo))	
	# arcpy.AddError(name)
	# arcpy.AddError(str(e))	
	# arcpy.AddError(str(tbinfo))	
                 
def write_log(pTipo, pEtapa, pCadena):
	folder_log = workspace + "\\LOGS"
	if not os.path.exists(folder_log):
		os.makedirs(folder_log)

	text_file = open(logs,'a')
	tiempo = datetime.now().strftime("%d/%m/%Y, %H:%M:%S - ")

	if(pTipo == 0):
		tipo = 'ERROR - '
	elif(pTipo == 1):
		tipo = 'INFO - '
		
	if(pEtapa == 0):
		etapa = '  EXECUTION - '
	elif(pEtapa == 1):
		etapa = '  BEGINNING - '
	elif(pEtapa == 2):
		etapa = '  END - '
	elif(pEtapa == 3):
		etapa = '  FUNCTION - '
	elif(pEtapa == 4):
		etapa = 'START - '
	elif(pEtapa == 5):
		etapa = 'CLOSING - '
	elif(pEtapa == 6):
		etapa = 'CONSULTATION - '
	elif(pEtapa == 7):
		etapa = '  CONECTION - '
		
	name = "%s%s%s%s"%(tipo,tiempo,etapa,pCadena)
	text_file.write(name+"\n")
	# arcpy.AddMessage(name)
	print(name)
	text_file.close()

def start_script():
    folder= workspace+"\\LOGS"
    if not os.path.exists(folder):
        os.makedirs(folder)
    text_file = open(logs,'a')
    text_file.write("\n"+(75*"=")+"\n")
    text_file.close()		
    name = "Script %s."%name_general
    write_log(1, 4, name)	
  
def end_script():
    # End script
    name = "Script %s."%name_general
    write_log(1, 5, name)
    
def updateShedule():
	name_f = "Upadete Sehedule"
	write_log(1,3,name_f)
	try:
		name = "1. Analyst data from Survey123 report."
		write_log(1,1,name)
		item 	= gis.content.get(id_survey)
		layer 	= item.layers[1]
		dflayer = pd.DataFrame.spatial.from_layer(layer)
		col_id 	= dflayer.columns[3]
		col_date = dflayer.columns[22]
		write_log(1,2,name)
  
		name = "2. Join item id field date."
		write_log(1,1,name)
		dflayer1 = dflayer.iloc[:,[3,22]]
		dflayer2 = dflayer.iloc[:,[5,22]]
		dflayer2 = dflayer2.rename(columns={dflayer2.columns[0]:col_id})
		dflayer3 = dflayer.iloc[:,[6,22]]
		dflayer3 = dflayer3.rename(columns={dflayer3.columns[0]:col_id})
		dfdate= pd.concat([dflayer1,dflayer2,dflayer3])
		write_log(1,2,name)
  
		name ="3. Analyst date report min and max group from id."
		write_log(1,1,name)
		dfdate = dfdate.groupby(col_id)
		dfdate = dfdate.agg(START_DATE_ACT=(col_date, np.min),DATE_ACT=(col_date, np.max))
		dfdate = dfdate.reset_index()
		dfdate = pd.DataFrame(dfdate)
		write_log(1,2,name)
  
		name = "4. Read Excel data."
		write_log(1,1,name)
		df = pd.read_excel(excel,sheet_name=0,header=1)
		df = df.iloc[:,:5]
		col_excel = df.columns[0]
		dfdate[col_id] 	= dfdate[col_id].astype("str")
		df[col_excel]	= df[col_excel].astype("str")
		write_log(1,2,name)
  
		name = "5. Join data Survey123 from excel data."
		write_log(1,1,name)
		dfanalyst = df.set_index(col_excel).join(dfdate.set_index(col_id))
		dfanalyst = dfanalyst.reset_index()
		col_date  = list(dfanalyst.columns)
		dfanalyst[col_date[2]] = pd.to_datetime(dfanalyst[col_date[2]])	
		dfanalyst[col_date[3]] = pd.to_datetime(dfanalyst[col_date[3]])	
		dfanalyst[col_date[5]] = pd.to_datetime(dfanalyst[col_date[5]])	
		dfanalyst[col_date[6]] = pd.to_datetime(dfanalyst[col_date[6]])	
		write_log(1,2,name)
  
		name = "6. Analyst data Calculate days report."
		write_log(1,1,name)
		# *=========================== Calculate days =============================================
		col_task 		= df.columns[1]
		col_start_date 	= df.columns[2]
		col_end_date 	= df.columns[3]
		col_status 		= df.columns[4]
		dfanalyst["DURATION"]	= (dfanalyst[col_end_date] - dfanalyst[col_start_date]).dt.days
		dfanalyst["EXEC_DAY"]	= (now - dfanalyst[col_start_date]).dt.days
		dfanalyst["DAY_REPORT"] = (dfanalyst["DATE_ACT"] - dfanalyst[col_start_date]).dt.days
		dfanalyst["DAY_ACT"] 	= (dfanalyst["START_DATE_ACT"] - dfanalyst[col_start_date]).dt.days
		dfanalyst["PERCENTAGE"] = round((dfanalyst["DAY_REPORT"] *100)/(dfanalyst["DURATION"]),2)
		# *==========================================================================================
		dfanalyst.loc [~ (dfanalyst ["EXEC_DAY"]> 0), "EXEC_DAY"] = np.nan
		dfanalyst.loc [~ (dfanalyst ["PERCENTAGE"]> 0), "PERCENTAGE"] = np.nan
		dfanalyst.loc [~ (dfanalyst ["PERCENTAGE"]< 100), "PERCENTAGE"] = np.nan
		dfanalyst.loc [~ (dfanalyst ["EXEC_DAY"]> 0), "EXEC_DAY"] = np.nan
		dfanalyst.fillna({'EXEC_DAY':0,'DAY_REPORT':0,'DAY_ACT':0,'PERCENTAGE':0}, inplace=True)
		write_log(1,2,name)
  
		name = "7. Validate field type from update services."
		write_log(1,1,name)
		colstr = list(dfanalyst.columns[0:6])
		colint = list(dfanalyst.columns[7:])
		dfanalyst[colstr] = dfanalyst[colstr].astype("str")
		dfanalyst[colint] = dfanalyst[colint].astype("int32")
		dfanalyst.fillna({	col_date[3]:'Nat',
                    		col_date[4]:'Nat',
                      		col_date[5]:'Nat',
                        	col_date[6]:'Nat'}, 
                   			inplace=True)
		write_log(1,2,name)
  
		name = "8. Update Arcgis Online layer."
		write_log(1,1,name)
		item_fc = gis.content.get(id_shedule)
		layer_fc= item_fc.layers[0]
		layer_fc = layer_fc.query()
		layer_fc = layer_fc.features
		features_for_update = []
		for ID in dfanalyst["ID"]:
			matching_row = dfanalyst.where(dfanalyst.ID==ID).dropna()
			original_feature  = [f for f in layer_fc if f.attributes["ID_SHEDULE"]==ID][0]
			feature_to_be_updated = deepcopy(original_feature)
			feature_to_be_updated.attributes['ACTIVITY'] 	= matching_row[col_task].values[0]
			feature_to_be_updated.attributes['START_DATE'] 	= matching_row[col_start_date].values[0]
			feature_to_be_updated.attributes['END_DATE'] 	= matching_row[col_end_date].values[0]
			feature_to_be_updated.attributes['DURATION'] 	= matching_row['DURATION'].values[0]
			feature_to_be_updated.attributes['EXEC_DAY'] 	= matching_row['EXEC_DAY'].values[0]
			feature_to_be_updated.attributes['DAY_REPORT']	= matching_row['DAY_REPORT'].values[0]
			feature_to_be_updated.attributes['DAY_ACT']		= matching_row['DAY_ACT'].values[0]
			if matching_row['START_DATE_ACT'].values[0] != "NaT":
				feature_to_be_updated.attributes['START_DATE_ACT']= matching_row['START_DATE_ACT'].values[0]
				feature_to_be_updated.attributes['DATE_ACT']= matching_row['DATE_ACT'].values[0]
				feature_to_be_updated.attributes['PERCENTAGE']= matching_row['PERCENTAGE'].values[0]
			if matching_row[col_status].values[0] != "nan":
				feature_to_be_updated.attributes['STATUS']= matching_row[col_status].values[0]
			features_for_update.append(feature_to_be_updated)
		flayer = item_fc.layers[0]
		flayer.edit_features(updates= features_for_update)
		write_log(1,2,name)
		name = "9. Update excel"
		write_log(1,1,name)
		wb = load_workbook(excel)
		wb.active = 0
		ws = wb.active
		count = 2
		for i, c in dfanalyst.iterrows():
			count +=1
			if c[5] != "NaT":
				ws['G%s'% count] = c[5]
				ws['H%s'% count] = c[6]
				ws['I%s'% count] = c[8]
				ws['J%s'% count] = c[9]
				ws['K%s'% count] = c[10]
				ws['L%s'% count] = c[11]
			else:
				ws['G%s'% count] = ""
				ws['H%s'% count] = ""
				ws['I%s'% count] = ""
				ws['J%s'% count] = ""
				ws['K%s'% count] = ""
				ws['L%s'% count] = ""
		wb.save(excel)
		write_log(1,2,name)
	except Exception as e:
		captureError(e,name)
	write_log(1,2,name_f)
 
# =============================== PROCESS ==================================
if __name__ == '__main__':
	start_script()
	try:
		open(excel,"r+")
		updateShedule()
	except IOError:
		name = "Excel is open..! please closed excel"
		# arcpy.AddError(name)
		print(name)
	end_script()

